#!/usr/bin/env python3
"""
Rust Skin Tracker
Reads your public Steam inventory for Rust and builds an Excel portfolio
sheet with current market values, purchase prices and P/L.

Not affiliated with Valve Corporation or Facepunch Studios.
"""

import json
import os
import re
import sys
import time
from datetime import datetime

try:
    import requests
except ImportError:
    print("Missing dependency: requests")
    print("Run:  pip install -r requirements.txt")
    sys.exit(1)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependency: openpyxl")
    print("Run:  pip install -r requirements.txt")
    sys.exit(1)


# ---------------------------------------------------------------- constants

RUST_APP_ID = 252490
CONFIG_FILE = "config.json"

# Steam currency codes -> (symbol, excel number format)
CURRENCIES = {
    1:  ("$",  '"$"#,##0.00'),
    2:  ("£",  '"£"#,##0.00'),
    3:  ("€",  '"€"#,##0.00'),
    5:  ("R$", '"R$"#,##0.00'),
    6:  ("zł", '#,##0.00" zł"'),
    7:  ("R",  '"R"#,##0.00'),
    20: ("CHF", '"CHF "#,##0.00'),
    23: ("C$", '"C$"#,##0.00'),
    24: ("A$", '"A$"#,##0.00'),
}

DEFAULT_CONFIG = {
    "steam_id": "",
    "currency": 3,
    "language": "en",
    "output_file": "rust_skin_tracker.xlsx",
    "request_delay": 1.2,
}

STRINGS = {
    "it": {
        "title": "RUST SKIN TRACKER",
        "updated": "Aggiornato il",
        "headers": ["NOME SKIN", "TIPO", "QTÀ", "SCAMBIABILE", "PREZZO ACQ.",
                    "VALORE ATT.", "DIFFERENZA", "VARIAZIONE (%)", "ULTIMO AGG."],
        "yes": "Sì", "no": "No", "na": "N/D",
        "total": "TOTALE PORTAFOGLIO",
        "summary_sheet": "Riepilogo",
        "inventory_sheet": "Inventario",
        "summary_title": "RIEPILOGO PORTAFOGLIO",
        "stats": ["Skin totali", "Skin scambiabili", "Con prezzo di mercato",
                  "Investimento totale", "Valore attuale", "P/L totale",
                  "Rendimento (%)", "Skin in profitto", "Skin in perdita"],
    },
    "en": {
        "title": "RUST SKIN TRACKER",
        "updated": "Updated on",
        "headers": ["SKIN NAME", "TYPE", "QTY", "TRADABLE", "BUY PRICE",
                    "CURRENT VALUE", "DIFFERENCE", "CHANGE (%)", "LAST UPDATE"],
        "yes": "Yes", "no": "No", "na": "N/A",
        "total": "PORTFOLIO TOTAL",
        "summary_sheet": "Summary",
        "inventory_sheet": "Inventory",
        "summary_title": "PORTFOLIO SUMMARY",
        "stats": ["Total skins", "Tradable skins", "With market price",
                  "Total invested", "Current value", "Total P/L",
                  "Return (%)", "Skins in profit", "Skins at a loss"],
    },
}

# Steam item type tags -> readable label
ITEMCLASS_MAP = {
    "it": {
        "shoes.boots": "Stivali", "pants": "Pantaloni", "hoodie": "Felpa",
        "jacket": "Giacca", "mask.bandana": "Bandana", "roadsign.jacket": "Giubbotto",
        "roadsign.kilt": "Kilt", "metal.facemask": "Maschera", "box.wooden": "Cassa",
        "sleepingbag": "Sacco a pelo", "hat.boonie": "Cappello", "hat.beenie": "Berretto",
        "hat.cap": "Cappellino", "burlap.shirt": "Camicia", "burlap.trousers": "Pantaloni",
        "burlap.shoes": "Scarpe", "burlap.gloves": "Guanti", "burlap.headwrap": "Copricapo",
        "tshirt": "Maglietta", "tshirt.long": "Maglietta lunga", "attire.hide.vest": "Gilet",
        "wood.armor.shirt": "Armatura legno", "wood.armor.pants": "Pantaloni legno",
        "roadsign.gloves": "Guanti", "coffeecan.helmet": "Elmetto", "rifle.ak": "AK-47",
        "rifle.bolt": "Bolt Action", "smg.mp5": "MP5", "pistol.python": "Revolver",
        "pistol.semiauto": "Pistola", "shotgun.pump": "Fucile a pompa", "lmg.m249": "M249",
        "door.hinged.metal": "Porta metallica", "door.double.hinged.metal": "Porta doppia",
        "furnace": "Fornace", "workbench1": "Banco lavoro 1",
        "workbench2": "Banco lavoro 2", "workbench3": "Banco lavoro 3",
    },
}

STEAMCAT_MAP = {
    "it": {
        "steamcat.clothing": "Abbigliamento", "steamcat.armor": "Armatura",
        "steamcat.weapon": "Arma", "steamcat.tool": "Attrezzo",
        "steamcat.construction": "Costruzione", "steamcat.misc": "Varie",
        "steamcat.deployable": "Oggetto", "steamcat.resources": "Risorse",
        "steamcat.medical": "Medico", "steamcat.food": "Cibo",
        "steamcat.ammunition": "Munizioni", "steamcat.traps": "Trappola",
        "steamcat.electrical": "Elettrico",
    },
}

# Colours
RUST_ORANGE, WHITE = "C44213", "FFFFFF"
RUST_GRAY, RUST_DARK, HEADER_BG = "2D2D2D", "252525", "1C1C1C"
GREEN, RED_COL = "4CAF50", "F44336"

_thin = Side(style="thin", color="444444")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

HEADERS = {"User-Agent": "rust-skin-tracker (github)"}


# ------------------------------------------------------------------- config

def load_config():
    cfg = dict(DEFAULT_CONFIG)
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, encoding="utf-8") as f:
                cfg.update(json.load(f))
        except (json.JSONDecodeError, OSError) as e:
            print(f"Warning: could not read {CONFIG_FILE} ({e}), using defaults.")

    if not cfg.get("steam_id"):
        cfg["steam_id"] = prompt_steam_id()
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(cfg, f, indent=2)
            print(f"  Saved to {CONFIG_FILE} — it won't ask again.\n")
        except OSError:
            pass

    if cfg["currency"] not in CURRENCIES:
        cfg["currency"] = 3
    if cfg["language"] not in STRINGS:
        cfg["language"] = "en"
    return cfg


def prompt_steam_id():
    print("Enter your SteamID64 (17 digits) or your profile URL.")
    print("Find it at https://steamid.io if you don't know it.\n")
    while True:
        raw = input("  SteamID64 / URL: ").strip()
        sid = extract_steam_id(raw)
        if sid:
            return sid
        if "/id/" in raw:
            print("  That's a custom URL. Paste it into https://steamid.io "
                  "and copy the steamID64 (17 digits).")
        else:
            print("  Not a valid SteamID64. It must be 17 digits.")


def extract_steam_id(raw):
    match = re.search(r"\b(7656\d{13})\b", raw)
    return match.group(1) if match else None


# -------------------------------------------------------------- steam calls

def http_get(url, params, retries=4):
    """GET with backoff on 429/5xx. Returns Response or None."""
    delay = 5
    for _ in range(retries):
        try:
            resp = requests.get(url, params=params, headers=HEADERS, timeout=20)
        except requests.RequestException:
            time.sleep(delay)
            delay *= 2
            continue
        if resp.status_code == 429 or resp.status_code >= 500:
            print(f"    Rate limited / server error, waiting {delay}s...")
            time.sleep(delay)
            delay *= 2
            continue
        return resp
    return None


def fetch_inventory(steam_id, delay):
    """Fetch the full inventory, following pagination."""
    url = f"https://steamcommunity.com/inventory/{steam_id}/{RUST_APP_ID}/2"
    assets, descriptions = [], []
    start_assetid, page = None, 1

    while True:
        params = {"l": "english", "count": 2000}
        if start_assetid:
            params["start_assetid"] = start_assetid

        print(f"  Fetching inventory (page {page})...")
        resp = http_get(url, params)

        if resp is None:
            raise RuntimeError(
                "Steam is rate limiting this IP. Wait a few hours and retry.")
        if resp.status_code in (401, 403):
            raise RuntimeError(
                "Inventory is private. Steam > Settings > Privacy > "
                "set 'Inventory' to Public.")
        if resp.status_code == 404:
            raise RuntimeError("Profile not found. Check your SteamID64.")
        if resp.status_code != 200:
            raise RuntimeError(f"Steam returned HTTP {resp.status_code}")

        try:
            data = resp.json()
        except ValueError:
            raise RuntimeError("Steam returned an unexpected response.")

        if not data or not data.get("assets"):
            if page == 1:
                raise RuntimeError("Inventory is empty or not accessible.")
            break

        assets.extend(data["assets"])
        descriptions.extend(data.get("descriptions", []))

        if data.get("more_items") and data.get("last_assetid"):
            start_assetid = data["last_assetid"]
            page += 1
            time.sleep(delay)
        else:
            break

    return {"assets": assets, "descriptions": descriptions}


def parse_price(text):
    """Parse Steam's localised price string into a float."""
    if not text:
        return None
    text = text.replace("\xa0", " ").replace("\u202f", " ")
    match = re.search(r"[\d][\d.,\s]*", text)
    if not match:
        return None
    num = match.group(0).replace(" ", "").strip(".,")

    if "." in num and "," in num:
        if num.rfind(",") > num.rfind("."):      # 1.234,56
            num = num.replace(".", "").replace(",", ".")
        else:                                     # 1,234.56
            num = num.replace(",", "")
    elif "," in num:
        head, _, tail = num.rpartition(",")
        num = f"{head}.{tail}" if len(tail) in (1, 2) else num.replace(",", "")
    else:
        head, _, tail = num.rpartition(".")
        if head and len(tail) not in (1, 2):
            num = num.replace(".", "")

    try:
        return float(num)
    except ValueError:
        return None


def get_market_price(market_hash_name, currency):
    url = "https://steamcommunity.com/market/priceoverview/"
    params = {"appid": RUST_APP_ID, "currency": currency,
              "market_hash_name": market_hash_name}
    resp = http_get(url, params, retries=3)
    if resp is None or resp.status_code != 200:
        return None
    try:
        data = resp.json()
    except ValueError:
        return None
    if not data.get("success"):
        return None
    return parse_price(data.get("lowest_price") or data.get("median_price"))


def parse_inventory(data, lang):
    descriptions = {}
    for d in data.get("descriptions", []):
        key = f"{d.get('classid')}_{d.get('instanceid', '0')}"
        descriptions[key] = d

    itemclass = ITEMCLASS_MAP.get(lang, {})
    steamcat = STEAMCAT_MAP.get(lang, {})

    skins = {}
    for asset in data.get("assets", []):
        key = f"{asset.get('classid')}_{asset.get('instanceid', '0')}"
        desc = descriptions.get(key)
        if not desc or not desc.get("marketable"):
            continue

        name = desc.get("market_hash_name") or desc.get("name")
        if not name:
            continue

        # Same skin owned multiple times: count it, don't drop it.
        if name in skins:
            skins[name]["qty"] += int(asset.get("amount", 1))
            continue

        tags = desc.get("tags", [])
        raw_class = next((t.get("internal_name", "") for t in tags
                          if t.get("category") == "itemclass"), "")
        raw_cat = next((t.get("internal_name", "") for t in tags
                        if t.get("category") == "steamcat"), "")
        label_cat = next((t.get("localized_tag_name", "") for t in tags
                          if t.get("category") == "steamcat"), "")

        item_type = (itemclass.get(raw_class)
                     or steamcat.get(raw_cat)
                     or label_cat or raw_cat or raw_class or "?")

        skins[name] = {
            "name": name,
            "type": item_type,
            "qty": int(asset.get("amount", 1)),
            "tradable": bool(desc.get("tradable")),
            "market_hash_name": desc.get("market_hash_name") or name,
        }

    return sorted(skins.values(), key=lambda s: s["name"])


# --------------------------------------------------------------------- xlsx

def load_existing_buy_prices(path, preferred_sheet):
    """Read buy prices from a previous run.

    Sheet names are localised, so a user who switches language would otherwise
    lose their prices. Try the current language first, then every other one.
    """
    if not os.path.exists(path):
        return {}
    try:
        wb = load_workbook(path)
    except (OSError, ValueError):
        return {}

    candidates = [preferred_sheet] + [s["inventory_sheet"] for s in STRINGS.values()]
    ws = next((wb[n] for n in candidates if n in wb.sheetnames), None)
    if ws is None:
        return {}

    prices = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if len(row) < 5:
            continue
        name, buy_price = row[0], row[4]   # A = name, E = buy price
        if name and isinstance(buy_price, (int, float)) and buy_price > 0:
            prices[str(name)] = float(buy_price)
    return prices


def ensure_writable(path):
    """Fail fast if the spreadsheet is locked (usually: open in Excel)."""
    if not os.path.exists(path):
        return
    try:
        with open(path, "r+b"):
            pass
    except PermissionError:
        raise RuntimeError(
            f"'{path}' is locked. Close it in Excel (or LibreOffice) and run again.")


def hstyle(cell, bg=RUST_ORANGE, fg=WHITE):
    cell.font = Font(bold=True, color=fg, name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def dstyle(cell, fg="CCCCCC", bg=RUST_GRAY, center=True, bold=False, fmt=None):
    cell.font = Font(color=fg, name="Arial", size=10, bold=bold)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center")
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt


def build_excel(skins, buy_prices, cfg):
    s = STRINGS[cfg["language"]]
    _, money = CURRENCIES[cfg["currency"]]
    pct = '+0.0%;-0.0%;"-"'
    alt_bg = [RUST_GRAY, RUST_DARK]

    wb = Workbook()
    ws = wb.active
    ws.title = s["inventory_sheet"]

    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value = s["title"]
    t.font = Font(bold=True, color=WHITE, name="Arial", size=16)
    t.fill = PatternFill("solid", start_color=RUST_ORANGE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:I2")
    sub = ws["A2"]
    sub.value = f"{s['updated']} {datetime.now():%d/%m/%Y %H:%M}"
    sub.font = Font(color="AAAAAA", name="Arial", size=9, italic=True)
    sub.fill = PatternFill("solid", start_color=HEADER_BG)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    for col, head in enumerate(s["headers"], 1):
        hstyle(ws.cell(3, col, head))
    ws.row_dimensions[3].height = 28

    last_row = 3
    for i, skin in enumerate(skins):
        r = last_row = i + 4
        bg = alt_bg[i % 2]
        bp = buy_prices.get(skin["name"], 0)
        cp = skin.get("price")

        dstyle(ws.cell(r, 1, skin["name"]), fg=WHITE, bg=bg, center=False, bold=True)
        dstyle(ws.cell(r, 2, skin["type"]), fg="AAAAAA", bg=bg)
        dstyle(ws.cell(r, 3, skin["qty"]), fg="CCCCCC", bg=bg)
        dstyle(ws.cell(r, 4, s["yes"] if skin["tradable"] else s["no"]),
               fg=GREEN if skin["tradable"] else "888888", bg=bg)
        dstyle(ws.cell(r, 5, bp or None), fg="4FC3F7", bg=bg, fmt=money)
        dstyle(ws.cell(r, 6, cp), fg="FFD54F", bg=bg, fmt=money)

        if cp is None:
            dstyle(ws.cell(r, 7, s["na"]), fg="666666", bg=bg)
            dstyle(ws.cell(r, 8, s["na"]), fg="666666", bg=bg)
        elif not bp:
            # No buy price yet: showing (value - 0) would fake a profit.
            dstyle(ws.cell(r, 7, "-"), fg="666666", bg=bg)
            dstyle(ws.cell(r, 8, "-"), fg="666666", bg=bg)
        else:
            colour = GREEN if (cp - bp) >= 0 else RED_COL
            dstyle(ws.cell(r, 7, f'=IF(E{r}="","-",(F{r}-E{r})*C{r})'),
                   fg=colour, bg=bg, bold=True, fmt=money)
            dstyle(ws.cell(r, 8, f'=IF(E{r}="","-",(F{r}-E{r})/E{r})'),
                   fg=colour, bg=bg, bold=True, fmt=pct)

        dstyle(ws.cell(r, 9, f"{datetime.now():%d/%m/%Y %H:%M}"), fg="888888", bg=bg)
        ws.row_dimensions[r].height = 22

    tot = last_row + 1
    ws.merge_cells(f"A{tot}:D{tot}")
    tc = ws[f"A{tot}"]
    tc.value = s["total"]
    tc.font = Font(bold=True, color=WHITE, name="Arial", size=11)
    tc.fill = PatternFill("solid", start_color=RUST_ORANGE)
    tc.alignment = Alignment(horizontal="right", vertical="center")
    tc.border = BORDER

    # Only rows with BOTH a buy price and a market price count toward P/L.
    both = f'(E4:E{last_row}<>"")*(F4:F{last_row}<>"")'
    pl = f'=SUMPRODUCT({both}*C4:C{last_row}*(F4:F{last_row}-E4:E{last_row}))'
    base = f'SUMPRODUCT({both}*C4:C{last_row}*E4:E{last_row})'
    totals = [
        (5, f"=SUMPRODUCT(C4:C{last_row},E4:E{last_row})", money),
        (6, f"=SUMPRODUCT(C4:C{last_row},F4:F{last_row})", money),
        (7, pl, money),
        (8, f'=IF({base}=0,"-",G{tot}/{base})', pct),
    ]
    for col, formula, fmt in totals:
        c = ws.cell(tot, col, formula)
        c.font = Font(bold=True, color=WHITE, name="Arial", size=11)
        c.fill = PatternFill("solid", start_color=RUST_ORANGE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        c.number_format = fmt
    ws.row_dimensions[tot].height = 26

    for i, w in enumerate([36, 15, 7, 13, 15, 15, 15, 14, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    # ---- summary sheet
    inv = s["inventory_sheet"]
    ws2 = wb.create_sheet(s["summary_sheet"])
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:B1")
    t2 = ws2["A1"]
    t2.value = s["summary_title"]
    t2.font = Font(bold=True, color=WHITE, name="Arial", size=14)
    t2.fill = PatternFill("solid", start_color=RUST_ORANGE)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32

    formulas = [
        f"=SUM('{inv}'!C4:C{last_row})",
        f"=SUMIF('{inv}'!D4:D{last_row},\"{s['yes']}\",'{inv}'!C4:C{last_row})",
        f"=COUNT('{inv}'!F4:F{last_row})",
        f"='{inv}'!E{tot}",
        f"='{inv}'!F{tot}",
        f"='{inv}'!G{tot}",
        f"='{inv}'!H{tot}",
        f"=COUNTIF('{inv}'!G4:G{last_row},\">0\")",
        f"=COUNTIF('{inv}'!G4:G{last_row},\"<0\")",
    ]
    fmts = [None, None, None, money, money, money, '+0.0%;-0.0%;0.0%', None, None]

    for i, (label, formula, fmt) in enumerate(zip(s["stats"], formulas, fmts)):
        r = i + 3
        lc = ws2.cell(r, 1, label)
        lc.font = Font(color="BBBBBB", name="Arial", size=10)
        lc.fill = PatternFill("solid", start_color=RUST_DARK)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lc.border = BORDER
        vc = ws2.cell(r, 2, formula)
        vc.font = Font(color=WHITE, name="Arial", size=10, bold=True)
        vc.fill = PatternFill("solid", start_color=RUST_GRAY)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = BORDER
        if fmt:
            vc.number_format = fmt
        ws2.row_dimensions[r].height = 22

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 20

    wb.save(cfg["output_file"])


# --------------------------------------------------------------------- main

def run():
    print("\n" + "=" * 55)
    print("  RUST SKIN TRACKER")
    print("=" * 55 + "\n")

    cfg = load_config()
    lang = cfg["language"]
    symbol, _ = CURRENCIES[cfg["currency"]]
    delay = float(cfg["request_delay"])

    ensure_writable(cfg["output_file"])

    print("[1/4] Loading Steam inventory...")
    data = fetch_inventory(cfg["steam_id"], delay)

    print("[2/4] Parsing marketable skins...")
    skins = parse_inventory(data, lang)
    if not skins:
        raise RuntimeError("No marketable Rust skins found in this inventory.")
    print(f"  Found {len(skins)} unique skins.")

    print("[3/4] Reading existing buy prices...")
    buy_prices = load_existing_buy_prices(
        cfg["output_file"], STRINGS[lang]["inventory_sheet"])
    print(f"  Restored {len(buy_prices)} buy prices.")

    print(f"[4/4] Fetching market prices (~{len(skins) * delay:.0f}s)...")
    for i, skin in enumerate(skins, 1):
        print(f"  [{i}/{len(skins)}] {skin['name'][:45]:<45}", end="\r")
        skin["price"] = get_market_price(skin["market_hash_name"], cfg["currency"])
        time.sleep(delay)
    print(" " * 70, end="\r")

    build_excel(skins, buy_prices, cfg)

    priced = [s for s in skins if s["price"] is not None]
    total = sum(s["price"] * s["qty"] for s in priced)
    items = sum(s["qty"] for s in skins)

    print("\n" + "=" * 55)
    print("  Done.")
    print(f"  Unique skins:    {len(skins)}  ({items} items)")
    print(f"  Priced:          {len(priced)}")
    print(f"  Portfolio value: {symbol}{total:.2f}")
    print(f"  File:            {cfg['output_file']}")
    print("=" * 55)
    print("\n  Enter your buy prices in column E — they are kept on re-runs.\n")


def main():
    try:
        run()
    except KeyboardInterrupt:
        print("\nCancelled.")
    except PermissionError as e:
        print(f"\nERROR: cannot write the file — is it open in Excel? ({e.filename})")
    except RuntimeError as e:
        print(f"\nERROR: {e}")
    except Exception as e:  # noqa: BLE001
        print(f"\nUNEXPECTED ERROR: {type(e).__name__}: {e}")

    if sys.stdin.isatty():
        input("\nPress Enter to exit...")


if __name__ == "__main__":
    main()
